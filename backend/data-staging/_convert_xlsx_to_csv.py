"""대학알리미 학과 마스터 xlsx → UTF-8 CSV 변환기.

사용:  python _convert_xlsx_to_csv.py
출력:  university_programs.csv (StagingDataLoader 가 읽는 표준 이름)
"""
import csv, os
import openpyxl

HERE = os.path.dirname(__file__)
XLSX = os.path.join(HERE, "학교별 교육편제단위 정보_20241007기준.xlsx")
CSV_OUT = os.path.join(HERE, "university_programs.csv")

# StagingDataLoader 가 기대하는 영문 헤더 ↔ 엑셀 한글 헤더 매핑
COLUMN_MAP = {
    "조사년도": "surveyYear",
    "조사차수": "surveyRound",
    "학교명": "universityName",
    "학교코드": "universityCode",
    "대학구분": "universityType",
    "본분교": "campusType",
    "학교구분": "schoolType",
    "(대학)지역": "region",
    "단과대학명": "collegeName",
    "단과대학코드": "collegeCode",
    "학교별학과코드": "departmentCode",
    "학부·과(전공)명": "departmentName",
    "주야간구분": "dayNight",
    "학과특성": "departmentTrait",
    "학과상태": "departmentStatus",
    "대계열분류": "fieldL1",
    "중계열분류": "fieldL2",
    "소계열분류": "fieldL3",
    "대학자체대계열": "universityField",
    "수업연한": "durationYears",
    "학위과정": "degreeType",
    "(학과)소재지": "deptRegion",
    "(학과)소재지(상세)": "deptRegionDetail",
}

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb["sheet0"]

# 헤더 행은 5행
header_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
out_headers = [COLUMN_MAP.get(h, h) for h in header_row]

written = 0
skipped_dropped = 0
with open(CSV_OUT, "w", encoding="utf-8", newline="") as fp:
    writer = csv.writer(fp)
    writer.writerow(out_headers)
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        # 폐지/모집중단 학과는 자동완성에서 잡음만 키우므로 제외
        status = row[14] if len(row) > 14 else None
        if isinstance(status, str) and ("폐지" in status or "모집중단" in status):
            skipped_dropped += 1
            continue
        writer.writerow([("" if v is None else v) for v in row])
        written += 1

print(f"wrote {CSV_OUT}")
print(f"rows written: {written}, dropped(폐지/모집중단): {skipped_dropped}")
